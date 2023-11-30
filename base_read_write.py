import os
import re

import phonenumbers
from openpyxl import load_workbook


def check_file(filename):
    if not os.path.exists(filename):
        return False

    return True


def is_valid(phone):
    phone = str(phone)
    phone = re.sub("[^0-9]", "", phone)
    phone = "+" + phone
    try:
        if len(phone) > 1:
            p = phonenumbers.is_valid_number(phonenumbers.parse(phone))
            return p
    except:
        return None


def base_read_write(*_, **kwargs):
    read = "read"
    write = "write"
    check = "check"

    filename = "base.xlsx"  # Имя файла

    if not check_file(filename):
        raise FileExistsError(f"Файл {filename} не найден!")

    wb = load_workbook(filename)
    ws = wb.active

    table = []

    for t in ws.iter_rows():
        row = [i.value for i in t]
        table.append(row)

    if "flag" in kwargs:
        if str(kwargs["flag"]).lower() == read:
            errors = 0
            for row in table:
                if errors > 1:
                    return "Номер не обнаружено!"
                if len(row) == 2:
                    phone = row[0]
                    name = row[1]

                    if str(phone).isalpha():
                        continue

                    return phone, name


                elif len(row) == 3:
                    phone = row[0]
                    name = row[1]
                    status = row[2]

                    if status is None and not str(phone).isalpha():
                        return phone, name

                if row == table[-1]:
                    return "EOF", "EOF"

        elif str(kwargs["flag"]).lower() == check:
            all_rows = []

            for row in table:
                index = table.index(row)
                phone = row[0]
                name = row[1]

                if is_valid(phone) is None:
                    continue

                if is_valid(phone) is False:
                    all_rows.append([phone, name, "Строка:" + str(index + 1)])

            return all_rows

        elif str(kwargs["flag"]).lower() == write:
            phone = kwargs["phone_number"] if "phone_number" in kwargs else None
            status = kwargs["status"] if "status" in kwargs else None

            for row in ws.iter_rows():
                phone2 = row[0].value
                if len(row) == 3:
                    if phone == phone2:
                        row[2].value = status
                        wb.save(filename)
                        return "Успешно записано!"
                elif len(row) == 2:
                    table2 = [r.value for r in row]
                    if phone == phone2:
                        if str(table[0][0]).isalpha():
                            ws[f'C{table2.index(phone) + 2}'] = status
                        else:
                            ws[f'C{table2.index(phone) + 1}'] = status

                        wb.save(filename)
                        return "Успешно записано!"
